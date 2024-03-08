import os
import pandas as pd
from openpyxl import Workbook
import progressbar

# Specify the input folder containing XLSX files
input_folder = 'Input/'

# Specify the output folder for the merged file
output_folder = 'Output/'

# List XLSX files in the input folder
xlsx_files = [file for file in os.listdir(input_folder) if file.endswith('.xlsx')]

# Create a new Excel workbook
result_workbook = Workbook()

# Initialize a progress bar
progress_bar_widgets = [
    ' [', progressbar.Percentage(), '] ',
    progressbar.Bar(), ' (', progressbar.ETA(), ') ',
]
progress_bar = progressbar.ProgressBar(widgets=progress_bar_widgets, max_value=len(xlsx_files)).start()

# Iterate through each XLSX file in the input folder and add its data to a new sheet
for file_index, file in enumerate(xlsx_files):
    # Update the progress bar
    progress_bar.update(file_index)

    # Read each sheet from the XLSX file into a DataFrame
    df = pd.read_excel(os.path.join(input_folder, file), sheet_name=None)

    # Iterate through each sheet and add its data to a new sheet in the result workbook
    for sheet_name, sheet_data in df.items():
        # Create a new sheet with the same name as the original sheet
        result_sheet = result_workbook.create_sheet(title=sheet_name)

        # Add the headers to the new sheet
        for col_num, value in enumerate(sheet_data.columns.values, 1):
            result_sheet.cell(row=1, column=col_num, value=value)

        # Add the data row by row to the new sheet
        for row_num, (_, row_data) in enumerate(sheet_data.iterrows(), 2):
            for col_num, value in enumerate(row_data, 1):
                result_sheet.cell(row=row_num, column=col_num, value=value)

# Remove the default sheet created by openpyxl
result_workbook.remove(result_workbook.active)

# Save the result workbook to a new XLSX file
result_workbook.save(os.path.join(output_folder, 'results.xlsx'))

# Finish the progress barp
progress_bar.finish()

print(f"Merge completed. Merged data saved to '{os.path.join(output_folder, 'results.xlsx')}'")
